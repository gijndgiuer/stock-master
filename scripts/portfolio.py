"""
股票持仓管理模块

功能：
- 创建 Excel 持仓模板
- 读取持仓数据
- 计算实时盈亏
- 记录交易历史

配置说明：
通过环境变量 PORTFOLIO_PATH 或 config.json 配置持仓文件路径
"""

from typing import Dict, List, Optional, Any
from datetime import datetime
from pathlib import Path
import json
import os


def get_portfolio_path(output_path: str = None) -> Path:
    """获取持仓文件路径"""
    if output_path:
        return Path(output_path)

    # 优先使用环境变量
    if os.environ.get('PORTFOLIO_PATH'):
        return Path(os.environ['PORTFOLIO_PATH'])

    # 其次使用配置文件
    config_path = Path('./config.json')
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            if config.get('portfolio_path'):
                return Path(config['portfolio_path'])

    # 默认使用当前目录
    return Path('./my_portfolio.xlsx')


def create_portfolio_template(output_path: str = None) -> str:
    """
    创建 Excel 持仓模板

    参数:
        output_path: 输出路径，默认为当前目录的 my_portfolio.xlsx

    返回:
        创建的文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    output_path = get_portfolio_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ===== Sheet 1: 持仓 =====
    ws_holdings = wb.active
    ws_holdings.title = "持仓"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 持仓表头
    holdings_headers = [
        "股票代码", "股票名称", "持仓数量", "买入均价",
        "当前价格", "市值", "盈亏金额", "盈亏比例", "买入日期", "备注"
    ]

    for col, header in enumerate(holdings_headers, 1):
        cell = ws_holdings.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 示例数据
    sample_holdings = [
        ["AAPL", "苹果公司", 10, 180.00, "", "", "", "", "2024-01-15", "长期持有"],
        ["TSLA", "特斯拉", 5, 400.00, "", "", "", "", "2024-06-01", ""],
        ["GOOGL", "谷歌", 3, 150.00, "", "", "", "", "2024-03-20", ""],
    ]

    for row_idx, row_data in enumerate(sample_holdings, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_holdings.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx >= 5:  # 当前价格及之后的列
                cell.alignment = Alignment(horizontal="right")

    # 添加公式说明
    ws_holdings.cell(row=6, column=1, value="说明：").font = Font(bold=True, color="FF0000")
    ws_holdings.cell(row=7, column=1, value="• 股票代码：美股用代码如 AAPL，港股用 0700.HK，A股用 600519.SS")
    ws_holdings.cell(row=8, column=1, value="• 当前价格、市值、盈亏：由 Claude 自动更新填写")
    ws_holdings.cell(row=9, column=1, value="• 只需填写：股票代码、持仓数量、买入均价、买入日期")

    # 调整列宽
    column_widths = [12, 15, 12, 12, 12, 15, 15, 12, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws_holdings.column_dimensions[get_column_letter(idx)].width = width

    # ===== Sheet 2: 交易记录 =====
    ws_transactions = wb.create_sheet("交易记录")

    transaction_headers = [
        "日期", "股票代码", "操作", "数量", "价格", "金额", "手续费", "备注"
    ]

    for col, header in enumerate(transaction_headers, 1):
        cell = ws_transactions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 示例交易
    sample_transactions = [
        ["2024-01-15", "AAPL", "买入", 10, 180.00, 1800.00, 0, "首次建仓"],
        ["2024-03-20", "GOOGL", "买入", 3, 150.00, 450.00, 0, ""],
        ["2024-06-01", "TSLA", "买入", 5, 400.00, 2000.00, 0, ""],
    ]

    for row_idx, row_data in enumerate(sample_transactions, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_transactions.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 调整列宽
    trans_widths = [12, 12, 8, 10, 12, 12, 10, 25]
    for idx, width in enumerate(trans_widths, 1):
        ws_transactions.column_dimensions[get_column_letter(idx)].width = width

    # ===== Sheet 3: 账户汇总 =====
    ws_summary = wb.create_sheet("账户汇总")

    summary_data = [
        ["账户汇总", ""],
        ["", ""],
        ["总投入", "=SUM(交易记录!F:F)"],
        ["当前市值", "（由 Claude 更新）"],
        ["总盈亏", "（由 Claude 更新）"],
        ["收益率", "（由 Claude 更新）"],
        ["", ""],
        ["最后更新时间", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]

    ws_summary.cell(row=1, column=1, value="账户汇总").font = Font(bold=True, size=14)

    for row_idx, (label, value) in enumerate(summary_data[2:], 3):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 20

    # ===== Sheet 4: 使用说明 =====
    ws_help = wb.create_sheet("使用说明")

    help_text = [
        ["股票持仓管理表 - 使用说明"],
        [""],
        ["一、持仓表（必填项）"],
        ["  • 股票代码：填写股票代码"],
        ["    - 美股：直接填代码，如 AAPL, TSLA, GOOGL"],
        ["    - 港股：代码后加 .HK，如 0700.HK, 9988.HK"],
        ["    - A股：代码后加 .SS(上海) 或 .SZ(深圳)，如 600519.SS"],
        ["  • 持仓数量：当前实际持有的股数"],
        ["  • 买入均价：当前持仓的平均成本价（做T后的实际均价）"],
        ["  • 买入日期：首次建仓日期 或 最近一次加仓日期"],
        [""],
        ["二、做T操作说明"],
        ["  • 持仓表只记录「结果」，不需要记录每次操作"],
        ["  • 例如：TSLA 做了几轮T，最终持仓100股，均价$380"],
        ["    → 直接填：数量=100，均价=380"],
        ["  • 均价可从券商App查看，或自己计算"],
        ["  • 想追踪每次操作盈亏，可在「交易记录」表记录（可选）"],
        [""],
        ["三、自动更新项"],
        ["  • 当前价格：Claude 会自动获取实时价格填入"],
        ["  • 市值 = 持仓数量 × 当前价格"],
        ["  • 盈亏金额 = 市值 - (持仓数量 × 买入均价)"],
        ["  • 盈亏比例 = 盈亏金额 / (持仓数量 × 买入均价) × 100%"],
        [""],
        ["四、交易记录（可选）"],
        ["  • 每次买入/卖出后，在交易记录表添加一行"],
        ["  • 操作类型：买入 / 卖出"],
        ["  • 不记录也不影响持仓分析"],
        [""],
        ["五、使用 Claude 分析"],
        ["  • 告诉 Claude：'分析我的持仓'"],
        ["  • Claude 会读取此表格，更新价格并给出建议"],
    ]

    for row_idx, row in enumerate(help_text, 1):
        ws_help.cell(row=row_idx, column=1, value=row[0] if row else "")

    ws_help.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_help.column_dimensions['A'].width = 60

    # 保存文件
    wb.save(output_path)
    print(f"✅ 持仓模板已创建: {output_path}")

    return str(output_path)


def read_portfolio(file_path: str = None) -> Dict[str, Any]:
    """
    读取 Excel 持仓文件

    参数:
        file_path: Excel 文件路径

    返回:
        包含持仓数据的字典
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    file_path = get_portfolio_path(file_path)
    wb = load_workbook(file_path, data_only=True)

    # 读取持仓表
    ws_holdings = wb["持仓"]

    holdings = []
    for row in ws_holdings.iter_rows(min_row=2, max_row=ws_holdings.max_row):
        # 跳过空行或说明行
        ticker = row[0].value
        if not ticker or ticker.startswith("说明") or ticker.startswith("•"):
            continue

        holding = {
            'ticker': str(ticker).upper(),
            'name': row[1].value or "",
            'shares': float(row[2].value or 0),
            'avg_cost': float(row[3].value or 0),
            'current_price': row[4].value,  # 可能为空
            'market_value': row[5].value,
            'profit_loss': row[6].value,
            'profit_loss_pct': row[7].value,
            'buy_date': str(row[8].value) if row[8].value else "",
            'notes': row[9].value or ""
        }

        if holding['shares'] > 0:
            holdings.append(holding)

    # 读取交易记录
    ws_trans = wb["交易记录"]

    transactions = []
    for row in ws_trans.iter_rows(min_row=2, max_row=ws_trans.max_row):
        date = row[0].value
        if not date:
            continue

        transaction = {
            'date': str(date),
            'ticker': str(row[1].value or "").upper(),
            'action': row[2].value or "",
            'shares': float(row[3].value or 0),
            'price': float(row[4].value or 0),
            'amount': float(row[5].value or 0),
            'fee': float(row[6].value or 0),
            'notes': row[7].value or ""
        }
        transactions.append(transaction)

    wb.close()

    return {
        'holdings': holdings,
        'transactions': transactions,
        'file_path': str(file_path),
        'read_time': datetime.now().isoformat()
    }


def update_portfolio_prices(file_path: str, prices: Dict[str, float]) -> Dict[str, Any]:
    """
    更新持仓表中的当前价格和盈亏

    参数:
        file_path: Excel 文件路径
        prices: 股票代码 -> 当前价格 的字典

    返回:
        更新后的持仓汇总
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    file_path = get_portfolio_path(file_path)
    wb = load_workbook(file_path)
    ws_holdings = wb["持仓"]

    total_cost = 0
    total_value = 0
    updated_holdings = []

    for row_idx in range(2, ws_holdings.max_row + 1):
        ticker = ws_holdings.cell(row=row_idx, column=1).value
        if not ticker or ticker.startswith("说明") or ticker.startswith("•"):
            continue

        ticker = str(ticker).upper()
        shares = float(ws_holdings.cell(row=row_idx, column=3).value or 0)
        avg_cost = float(ws_holdings.cell(row=row_idx, column=4).value or 0)

        if shares <= 0:
            continue

        # 获取当前价格
        current_price = prices.get(ticker)
        if current_price:
            # 计算市值和盈亏
            cost_basis = shares * avg_cost
            market_value = shares * current_price
            profit_loss = market_value - cost_basis
            profit_loss_pct = (profit_loss / cost_basis * 100) if cost_basis > 0 else 0

            # 更新单元格
            ws_holdings.cell(row=row_idx, column=5, value=round(current_price, 2))
            ws_holdings.cell(row=row_idx, column=6, value=round(market_value, 2))
            ws_holdings.cell(row=row_idx, column=7, value=round(profit_loss, 2))
            ws_holdings.cell(row=row_idx, column=8, value=f"{profit_loss_pct:.2f}%")

            total_cost += cost_basis
            total_value += market_value

            updated_holdings.append({
                'ticker': ticker,
                'shares': shares,
                'avg_cost': avg_cost,
                'current_price': current_price,
                'market_value': market_value,
                'profit_loss': profit_loss,
                'profit_loss_pct': profit_loss_pct
            })

    # 更新汇总表
    ws_summary = wb["账户汇总"]
    ws_summary.cell(row=4, column=2, value=round(total_value, 2))
    ws_summary.cell(row=5, column=2, value=round(total_value - total_cost, 2))
    total_return = ((total_value - total_cost) / total_cost * 100) if total_cost > 0 else 0
    ws_summary.cell(row=6, column=2, value=f"{total_return:.2f}%")
    ws_summary.cell(row=8, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    wb.save(file_path)
    wb.close()

    return {
        'total_cost': total_cost,
        'total_value': total_value,
        'total_profit_loss': total_value - total_cost,
        'total_return_pct': total_return,
        'holdings': updated_holdings,
        'update_time': datetime.now().isoformat()
    }


def format_portfolio_summary(portfolio_data: Dict[str, Any]) -> str:
    """
    格式化持仓汇总为易读文本
    """
    holdings = portfolio_data.get('holdings', [])

    if not holdings:
        return "持仓为空，请先添加股票"

    lines = [
        "**持仓汇总**",
        "",
        "| 股票 | 数量 | 成本价 | 现价 | 盈亏 |",
        "|------|------|--------|------|------|"
    ]

    for h in holdings:
        profit_emoji = "+" if h.get('profit_loss', 0) >= 0 else ""
        lines.append(
            f"| {h['ticker']} | {h['shares']:.0f} | ${h['avg_cost']:.2f} | "
            f"${h.get('current_price', 0):.2f} | {profit_emoji}{h.get('profit_loss_pct', 0):.1f}% |"
        )

    total = portfolio_data.get('total_profit_loss', 0)
    total_pct = portfolio_data.get('total_return_pct', 0)
    total_sign = "+" if total >= 0 else ""

    lines.extend([
        "",
        f"**总投入**: ${portfolio_data.get('total_cost', 0):,.2f}",
        f"**当前市值**: ${portfolio_data.get('total_value', 0):,.2f}",
        f"**总盈亏**: {total_sign}${total:,.2f} ({total_sign}{total_pct:.2f}%)"
    ])

    return "\n".join(lines)


if __name__ == "__main__":
    # 测试创建模板
    path = create_portfolio_template()
    print(f"\n测试读取: {path}")
    data = read_portfolio(path)
    print(f"读取到 {len(data['holdings'])} 条持仓记录")
